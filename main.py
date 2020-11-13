import openpyxl
import string
import numbers


def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def get_value(input):
    if isinstance(input.value, numbers.Integral):
        return input.value
    else:
        return 0


def calcMax(input):
    if isinstance(input[0].value, numbers.Integral):
        maximum = input[0].value
    else:
        print('ohno')
        maximum = 0.0

    for x in input:
        x_value = get_value(x)
        if x_value > maximum:
            maximum = x_value

    return maximum


def calcMin(input):
    if isinstance(input[0].value, numbers.Integral):
        minimum = input[0].value
    else:
        print('ohno')
        minimum = 0.0

    for x in input:
        x_value = get_value(x)
        if x_value < minimum:
            minimum = x_value

    return minimum


work_book = openpyxl.load_workbook('input3.xlsx', data_only=True)
data_sheet = work_book.get_sheet_by_name('TIM')

print('number of rows is ' + str(data_sheet.max_row) + '\n')

i = 7
while(i <= data_sheet.max_row):
    print('processing row: ' + str(i)+'\n')
    if data_sheet.cell(row=i, column=col2num('AF')).value is not None:
        print('condition met at row: ' + str(i)+'\n')
        #  potom v tretom riadku nad tymto riadkom (riadok 14) ma byt: 0,9*(max(AA14 az AA16)- min(AA14 az AA16)
        aaMax = calcMax(data_sheet['AA'][i - 3:i])
        aaMin = calcMin(data_sheet['AA'][i - 3:i])
        data_sheet.cell(row=i-3, column=col2num('AF')).value = 0.9 * (aaMax - aaMin)

        # V riadku druhom nad tymto riadkom (riadok 15) ma byt: 0,5*(max(AA14 az AA16)-min(AA14 az AA16)
        data_sheet.cell(row=i-2, column=col2num('AF')).value = 0.5 * (aaMax - aaMin)

        # V riadku nad tymto riadkom (riadok 16) ma byt: 0,1*(max (AA14 az AA16)-min(AA14 az AA16)
        data_sheet.cell(row=i-1, column=col2num('AF')).value = 0.1 * (aaMax - aaMin)

        if i+3 <= data_sheet.max_row:
            aa2Max = calcMax(data_sheet['AA'][i+1:i+3])
            aa2Min = calcMin(data_sheet['AA'][i+1:i+3])

            # V riadku pod tymto riadkom (AA18) ma byt: 0,1*(max(AA18 az AA20)-min(AA18 az AA20)
            data_sheet.cell(row=i+1, column=col2num('AF')).value = 0.1 * (aa2Max - aa2Min)

            # V riadku o dva nizsie (riadok AA19) ma byt: 0,5*(max(AA18 az AA20)-min(AA18 az AA20)
            data_sheet.cell(row=i+2, column=col2num('AF')).value = 0.5 * (aa2Max - aa2Min)

            # V riadku o tri nizsie (riadok AA20) ma byt: 0,9*(max(AA18 az AA20)- min (AA18 az AA20)
            data_sheet.cell(row=i+3, column=col2num('AF')).value = 0.9 * (aa2Max - aa2Min)
        i = i+4
    else:
        i = i+1

work_book.save('output2.xlsx')
